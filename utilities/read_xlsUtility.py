import openpyxl

class XlsUtility:

    @staticmethod
    def open_excel_file(excel_file,sheet_name):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        return workbook,sheet

    @staticmethod
    def get_max_rows(excel_file,sheet_name):
        workbook,sheet=XlsUtility.open_excel_file(excel_file,sheet_name)
        max_rows=sheet.max_row
        workbook.close()
        return max_rows


    @staticmethod
    def read_excel_file(excel_file,sheet_name,row_no,col_no):
        workbook, sheet = XlsUtility.open_excel_file(excel_file, sheet_name)
        data = sheet.cell(row_no,col_no).value
        workbook.close()
        return data

    @staticmethod
    def write_excel_file(excel_file, sheet_name, row_no, col_no,data):
        workbook, sheet = XlsUtility.open_excel_file(excel_file, sheet_name)
        sheet.cell(row_no, col_no).value=data
        workbook.save(excel_file)
        workbook.close()



